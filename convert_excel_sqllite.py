# -*- coding: utf-8 -*-
import sys, os
import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import *
from openpyxl.styles import *
from openpyxl.styles.colors import Color
#from openpyxl.drawing import *
from dateutil import parser
from common.webdatastore import webdatastore
objdb = webdatastore()
import DBRS_Currency_Frmt
reload(sys)
sys.setdefaultencoding('utf8')

from datetime import *

def disableprint():
    sys.stdout = open(os.devnull, 'w')
    pass

def enableprint():
    sys.stdout = sys.__stdout__

class xlsxReaderAdvance(object):
    def __init__(self):
        self.conObj = {} 

    def get_mergecells_dict(self, merged_cells):
        mdict = {}
        for mcell in merged_cells:
            rcells = range_boundaries(mcell)
            mdict[(rcells[1], rcells[0])] = (rcells[3], rcells[2])
            
        return mdict

    def get_cdict(self):
        cdict = {'data' :'',
                'formula': '',
                'formatStr': '',
                'comment_dict':{},
                'colspan':1,
                'rowspan':1,
                'font_dict': {},
                'border_dict':{},
                'fill_dict':{},
                'align_dict':{}
                } 
        return cdict

    def getColorDict(self, colrObj):
        if not colrObj:
            return {}
        
        colordict = {}
        colordict['tint'] = colrObj.tint
        if type( colrObj.auto) == bool:
            colordict['auto'] = colrObj.auto
        else:
            colordict['auto'] = False
        if type(colrObj.theme) == long:
            colordict['theme'] = colrObj.theme
        else:
            colordict['theme'] = None
        if type(colrObj.rgb) == str:
            clr = colrObj.rgb
            clr = clr[2:]+clr[:2]
            clr = 'rgb'+str(tuple(int(clr[i:i+2], 16) for i in (0, 2 ,4)))
            colordict['rgb'] = clr
        else:
            colordict['rgb'] = None
        if type(colrObj.indexed) == long:
            colordict['indexed'] = colrObj.indexed
        else:
            colordict['indexed'] = None
        colordict['type'] = colrObj.type
        return colordict

    def getFontDict(self, fontObj):
        if not fontObj: return {}
        fontdict = {}
        fontdict['name'] = fontObj.name
        fontdict['charset'] = fontObj.charset
        fontdict['family'] = fontObj.family
        fontdict['b'] = fontObj.b
        fontdict['i'] = fontObj.i
        fontdict['u'] = fontObj.u
        fontdict['strike'] = fontObj.strike
        fontdict['outline'] = fontObj.outline
        fontdict['shadow'] = fontObj.shadow
        fontdict['condense'] = fontObj.condense
        fontdict['extend'] = fontObj.extend
        fontdict['sz'] = fontObj.sz
        fontdict['vertAlign'] = fontObj.vertAlign
        fontdict['scheme'] = fontObj.scheme
        fontdict['colordict'] = self.getColorDict(fontObj.color)
        return fontdict

    def getBorderStyleDict(self, styleObj):
        if not styleObj: return {}
        styledict = {}
        colordict = self.getColorDict(styleObj.color)
        styledict['colordict'] = colordict
        styledict['style'] = styleObj.style        
        return styledict


    def getBorderDict(self, borderObj):
        if not borderObj: return {}
        border_dict = {}
        border_dict['outline'] = borderObj.outline
        border_dict['diagonalUp'] = borderObj.diagonalUp
        border_dict['diagonalDown'] = borderObj.diagonalDown
        border_dict['left'] = self.getBorderStyleDict(borderObj.left)
        border_dict['right'] = self.getBorderStyleDict(borderObj.right)
        border_dict['top'] = self.getBorderStyleDict(borderObj.top)
        border_dict['bottom'] = self.getBorderStyleDict(borderObj.bottom)
        border_dict['diagonal'] = self.getBorderStyleDict(borderObj.diagonal)
        border_dict['vertical'] = borderObj.vertical
        border_dict['horizontal'] = borderObj.horizontal
        return border_dict

    def getFillDict(self, fillObj):
        if not fillObj: return {}
        fill_dict = {}
        fill_dict['patternType'] = fillObj.patternType
        fill_dict['fgColor'] = self.getColorDict(fillObj.fgColor)
        fill_dict['bgColor'] = self.getColorDict(fillObj.bgColor)
        return fill_dict

    def getAlignmentDict(self, alignObj):
        if not alignObj: return {}
        align_dict = {}
        align_dict['indent'] = alignObj.indent
        align_dict['vertical'] = alignObj.vertical
        align_dict['relativeIndent'] = alignObj.relativeIndent
        align_dict['shrinkToFit'] = alignObj.shrinkToFit
        align_dict['justifyLastLine'] = alignObj.justifyLastLine
        align_dict['readingOrder'] = alignObj.readingOrder
        align_dict['wrapText'] = alignObj.wrapText
        align_dict['horizontal'] = alignObj.horizontal
        align_dict['textRotation'] = alignObj.textRotation
        return align_dict

    def getCommentDict(self, commentObj):
        comment_dict = {}
        comment_dict['author'] = commentObj.author
        #comment_dict['bind'] = commentObj.bind
        comment_dict['content'] = commentObj.content
        comment_dict['height'] = commentObj.height
        #comment_dict['parent'] = commentObj.parent
        comment_dict['text'] = commentObj.text
        #comment_dict['unbind'] = commentObj.unbind
        comment_dict['width'] = commentObj.width
        return comment_dict

    def getGeographicalPropDict(self, geoPropObj):
        geoprop_dict = {}
         

        return geoprop_dict

    def getChartDict(self, chartObj):
        chart_dict = {}

        return chart_dict

    def readExcel(self, fname, csv_file_path, data_flg, comment_flg):
        shnames_info = []
        
        excel_data = {} 
        data_wb = load_workbook(filename=fname, data_only=data_flg)
        wb = load_workbook(filename=fname, data_only=comment_flg)
        visible_sheets = [idx for idx, sheet in enumerate(wb._sheets) if sheet.sheet_state == "visible"]
        sheet_names = wb.get_sheet_names()
        active_sheet = wb.active.title
        for idx, sheet_name in enumerate(sheet_names):
            sheetObj = wb.get_sheet_by_name(sheet_name)

            print '---------------------------------------------------'
            print sheet_name
            print dir(sheetObj)

            #ROW HEIGHT
            row_outline_dict, rowHeights_dict = {}, {}
            for i in range(sheetObj.max_row):
                ob = sheetObj.row_dimensions[i+1]
                if ob.height: rowHeights_dict[i+1] = ob.height
                if ob.outline_level: 
                    if not row_outline_dict.get(i+1):
                        row_outline_dict[i+1] = []
                    row_outline_dict[i+1].append({'outline_level': ob.outline_level, 'hidden': ob.hidden})

            #COLUMN WIDTH
            col_outline_dict, colWidths_dict = {}, {}
            columns = []
            for i in range(sheetObj.max_column):
                ob = sheetObj.column_dimensions[get_column_letter(i+1)]
                columns.append(get_column_letter(i+1))
                if ob.width: colWidths_dict[get_column_letter(i+1)] = ob.width
                if ob.outline_level: 
                    if not col_outline_dict.get(i+1):
                        col_outline_dict[i+1] = []
                    col_outline_dict[i+1].append({'outline_level':ob.outline_level, 'hidden': ob.hidden})
                    
            #GRIDLINES
            gridlines = sheetObj.sheet_view.showGridLines
          
            #FREEZE_PANES 
            freeze_panes    = sheetObj.freeze_panes
            ac              = sheetObj.views.sheetView[0].selection[0].activeCell
            activeCell      = {'v1':ac, 'v2':coordinate_to_tuple(ac)}
            
            merged_cell_ranges = sheetObj.merged_cell_ranges
            mdict = self.get_mergecells_dict(merged_cell_ranges)
            sheet_name = str(sheet_name)
            excel_data[sheet_name] = {}
            excel_data[sheet_name]['visible'] = 1 if idx in visible_sheets else 0
            excel_data[sheet_name]['sheet_order'] = idx
            excel_data[sheet_name]['column_widths_dict'] = colWidths_dict
            excel_data[sheet_name]['row_heights_dict'] = rowHeights_dict
            excel_data[sheet_name]['gridlines'] = gridlines
            excel_data[sheet_name]['freeze_panes'] = freeze_panes
            excel_data[sheet_name]['merged_cell_ranges'] = merged_cell_ranges
            excel_data[sheet_name]['col_outline_dict'] = col_outline_dict
            excel_data[sheet_name]['row_outline_dict'] = row_outline_dict
            excel_data[sheet_name]['columns'] = columns
            excel_data[sheet_name]['activeCell'] = activeCell
            excel_data[sheet_name]['sheet_name'] = sheet_name
            celldict = {}
            mrow, mcol = 0, 0
            for rowid, rowObjs in enumerate(sheetObj.rows, 1):
                mrow = max(mrow, rowid)
                #dataSheetObj_rows = list(dataSheetObj.rows) 
                for colidx, cellObj in enumerate(rowObjs):
                    colid = colidx + 1
                    mcol = max(mcol, colid)
                    print cellObj
                    #print dir(cellObj)  
                    formula = ''
                    try:
                        formula = cellObj.value
                    except:
                        print 'ERROR IN VAL - ', sheet_name, rowid, colid, cellObj 
                    print sheet_name, rowid, colid, cellObj.value
                    formula = cellObj.value
                    print sheet_name, "=====", rowid, colid, formula
                    if str(formula) and str(formula)[0] != '=': formula = ''
                    formatStr = cellObj.number_format
                    formatStr = formatStr.replace('\\', '') 
                    font_dict, border_dict, fill_dict, align_dict = {}, {}, {}, {}
                    comment_dict = {}
                    if cellObj.has_style:
                        font_dict = self.getFontDict(cellObj.font)
                        border_dict = self.getBorderDict(cellObj.border)
                        fill_dict = self.getFillDict(cellObj.fill)
                        align_dict = self.getAlignmentDict(cellObj.alignment)
                        print '>>>> ', font_dict, border_dict, fill_dict, align_dict
                    if cellObj.comment:
                        comment_dict = self.getCommentDict(cellObj.comment)
                        
                    data = ''
                    cell = str((rowid, colid))
                    cdict = self.get_cdict()
                    cdict['formula'] = formula
                    cdict['comment_dict'] = comment_dict
                    cdict['formatStr'] = formatStr
                    cdict['font_dict'] = font_dict
                    cdict['border_dict'] = border_dict
                    cdict['fill_dict'] = fill_dict
                    cdict['align_dict'] = align_dict
                    if mdict.get(cell, ()):
                        rspan = mdict[cell][0] - rowid
                        cspan = mdict[cell][1] - colid
                        cdict['rowspan'] = rspan + 1
                        cdict['colspan'] = cspan + 1

                    celldict[cell] = cdict
                excel_data[sheet_name]['celldict'] = celldict
                excel_data[sheet_name]['rows'] = mrow
                excel_data[sheet_name]['cols'] = mcol
            '''
            wb_active = False
            if(active_sheet == sheet_name):
                wb_active = True
            shnames_info.append({'sheet_order':idx, 'sheet_name':sheet_name, 'active':wb_active})
            s_path = os.path.join(csv_file_path, 'sheet_info', 'filedata.dat')
            objdb.write_to_shelve(s_path, shnames_info)
            '''
            

        #Adding data, it include calculated data
        for idx, sheet_name in enumerate(sheet_names):
            dataSheetObj = data_wb.get_sheet_by_name(sheet_name)
            sheet_name = str(sheet_name)
            for rowid, rowObjs in enumerate(dataSheetObj.rows, 1):
                for colidx, cellObj in enumerate(rowObjs):
                    colid = colidx + 1

                    cell = str((rowid, colid))
                    data =  cellObj.value
                    if excel_data[sheet_name]['celldict'].get(cell, {})['formatStr']:
                        number_format = excel_data[sheet_name]['celldict'].get(cell, {})['formatStr']
                        if data and number_format!= 'General':
                            data = self.__get_formatted_text(data, number_format)

                    if data == None : data = ''
                    try: data = str(data)
                    except: pass
                    data = ' '.join(map(lambda data:data.strip(), data.split()))

                    if not excel_data[sheet_name]['celldict'].get(cell, {}):
                        excel_data[sheet_name]['celldict'][cell] = self.get_cdict()
                            
                    excel_data[sheet_name]['celldict'][cell]['data'] = data
            '''
            fin_dict = excel_data[sheet_name]
            s_path = os.path.join(csv_file_path, 'sheet_name', sheet_name)
            objdb.write_to_shelve(os.path.join(s_path, 'filedata.dat'), fin_dict)
            '''

        return  excel_data
        #return True

    def __get_formatted_text(self, value, fmt_str):
        format_str = fmt_str.replace("\\",'')
        if 'mm' in format_str.lower() or 'dd' in format_str.lower() or 'yyyy' in format_str.lower() or 'yy' in format_str.lower() or 'm' in format_str.lower() or 'd' in format_str.lower() or 'y' in format_str.lower():
            dt = parser.parse(value)
            res_date = self.get_date(dt.year, dt.month, dt.day, fmt_str)
            return res_date
        elif format_str:
            s_f,color = DBRS_Currency_Frmt.extract_currency_old('', format_str, value)
            if s_f:
                return s_f
        return value

    def get_date(self, year, month, day, format_key):
        year = str(year)
        month = str(month)
        day = str(day)
        identifier = ""
        res_arr = []
        for key in format_key:
            key = key.lower()
            if key not in ['m','y','d']:
                identifier = key
                break
        if identifier:
            format_arr = format_key.split(identifier)
            for each in format_arr:
                each = ''.join(e for e in each if e.isalnum())
                if each.lower() in ['mm', 'm']:
                    if (len(each) == 2) and len(month) == 1:
                        month = '0%s' %month
                    res_arr.append(month)
                elif each.lower() in ['dd', 'd']:
                    if (len(each) == 2) and len(day) == 1:
                        day = '0%s' %day
                    res_arr.append(day)
                elif each.lower() in ['yy','yyyy']:
                    if (len(each) == 2) and len(year) == 4:
                         year = year[2:]
                    elif (len(each) == 4) and len(year) == 2:
                         year = "20%s" % year
                    res_arr.append(year)
        res_date = identifier.join(res_arr)
        return res_date

    def process(self, dataPath, file_name, csv_file_path, data_flg=True, comment_flg=False, toprint=False):
        if toprint:
            enableprint()
        iFile = os.path.join(dataPath, file_name)
        if not os.path.exists(iFile): return {}
        if os.path.exists(csv_file_path):
            cmd = 'rm -rf %s && mkdir -p %s' %(csv_file_path, csv_file_path)
            os.system(cmd)
        else:
            cmd = 'mkdir -p %s' %(csv_file_path)
            os.system(cmd)
        print 'Reading file : ', iFile
        excel_op_data = self.readExcel(iFile, csv_file_path, data_flg, comment_flg)
        return excel_op_data

    def read_excel_save_sqlite(self, efname, epath, spath, sfname):
        ret = self.readExcel(efname, epath, True, False)
        enableprint()
        import create_db_sqllite_newvalidation_ashwat as create_db_sqllite_newvalidation
        cobj = create_db_sqllite_newvalidation.create_db_sqllite_newvalidation()
        path = spath
        dbname = sfname
        first = 1
        for k, vdict in ret.items():
            k = k.replace('-','')
            k = k.replace(' ','')
            rows = vdict.get('rows')
            cols = vdict.get('cols')
            if not rows:continue
            if not cols:continue
            #print k, rows, "====", cols
            #print vdict.get('celldict').keys()
            headerdict = {}
            header = []
            all_data = {} 
            for r in range(1,rows+1):
                data_dict = {}
                for c in range(1,cols+1):
                    rc = str((r,c))
                    rcdata = vdict.get('celldict').get(rc).get('data')
                    #print r, "=====", c, "======", rcdata
                    if r == 1:
                        if rcdata=='':continue            
                        if rcdata == 'INDEX':
                            rcdata = 'CINDEX'
                        rcdata = rcdata.replace('-','')
                        rcdata = rcdata.replace(' ','')
                        headerdict[c] = rcdata
                        print r, "=====", c, "======", rcdata
                        header.append(rcdata)
                    else:
                        #print c, "=====", headerdict
                        key = headerdict.get(c)
                        if not key:continue
                        data_dict[key] = rcdata
                if r != 1:
                    all_data[r] = copy.deepcopy(data_dict)
            '''
            print "========================"
            print headerdict
            print "========================"
            for k, v in all_data.items():
                print "*****************", k
                print v
            '''
            cobj.main_excel_data(path, dbname, k, header, first)
            kdata = []
            for r in sorted(all_data.keys()):
                rdict = all_data[r]
                rlst = []
                for key in header:
                    kcdata  = rdict.get(key, '')
                    rlst.append(kcdata)
                kdata.append(tuple(rlst))
            conn = cobj.get_sqllite_conn(path, dbname)
            cobj.insert_excel_data(conn, k, header,  kdata)
            first = 0
            #break

    def read_sqlite_to_excel(self, ifilename, ofilename):
        import sqlite3
        from xlsxwriter.workbook import Workbook

        workbook = Workbook(ofilename)

        conn=sqlite3.connect(ifilename)
        cur=conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
        res =  cur.fetchall()
        newline_indent = '\n   '
        for r in res:
            table_name = str(r[0])
            worksheet = workbook.add_worksheet(table_name)
            print "====================================================", table_name
            result = cur.execute("PRAGMA table_info('%s')" % table_name).fetchall()
            column_names = zip(*result)[1]
            r = 1 
            c = -1
            for column in column_names:
                c = c + 1
                worksheet.write(r, c, column)
            sql = "select * from %s" %(table_name)
            print sql
            mysel=cur.execute(sql)
            for i, row in enumerate(mysel):
                r = r + 1
                c = -1
                for j, value in enumerate(row):
                    c = c + 1
                    worksheet.write(r, c, value)
            #print ("\ncolumn names for %s:" % table_name)+newline_indent+(newline_indent.join(column_names))
        conn.close()
        workbook.close()

if __name__=="__main__":
    #disableprint()
    obj = xlsxReaderAdvance()
    #dataPath = sys.argv[1]
    #caseID = sys.argv[2] 
    #obj.process(dataPath, caseID)
    #obj.process('/var/www/html/WorkSpaceBuilder_DB/67/1/upload', 'BankRatiosTaxonomy_v4.xlsx', '/var/www/html/WorkSpaceBuilder_DB/67/1/xdata/docs/1', True, False, True)
    #ret = obj.read_excel_save_sqlite('BankRatiosTaxonomy_v4.xlsx', '/root/databuilder_train_ui/tenkTraining/Data_Builder_Training_Copy/pysrc/', '/tmp/', 'taxo_data')
    #ret = obj.read_excel_save_sqlite('FE-ReviewLinkbase1.xlsx', '/root/databuilder_train_ui/tenkTraining/Data_Builder_Training_Copy/pysrc/', '/tmp/', 'linkbase_data1')
    ret = obj.read_sqlite_to_excel('/tmp/linkbase_data1.db', '/var/www/html/output2.xlsx')

